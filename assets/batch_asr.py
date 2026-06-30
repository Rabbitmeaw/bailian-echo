#!/usr/bin/env python3
"""
批量视频 ASR 转写工具 — 基于阿里云百炼 CLI (bl speech recognize / fun-asr)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

遍历指定文件夹中的视频文件，逐一转写为文字，
输出 Excel (.xlsx) 或 CSV 文件。

用法:
    python3 batch_asr.py --folder /path/to/videos
    python3 batch_asr.py --folder /path/to/videos --format csv
    python3 batch_asr.py --folder /path/to/videos --output /path/to/result.xlsx

依赖:
    - bl (bailian-cli) 已安装并已鉴权
    - Python 3.8+ / openpyxl
"""

import argparse
import json
import os
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

VIDEO_EXTENSIONS = {
    '.mp4', '.mov', '.mkv', '.avi', '.webm', '.flv',
    '.wmv', '.m4v', '.ts', '.mts', '.m2ts', '.3gp',
    '.ogv', '.mpg', '.mpeg', '.rmvb', '.asf', '.vob',
}

ASR_TIMEOUT = 900

FIELD_NAMES = [
    '文件名', '文件路径', '时长(秒)', '文件大小(MB)',
    '完整文本', '处理状态', '处理耗时(秒)', '错误信息',
]


def get_file_size_mb(filepath: str) -> float:
    return os.path.getsize(filepath) / (1024 * 1024)


def run_asr(filepath: str) -> tuple[str | None, float | None, str | None]:
    """调用 bl speech recognize，返回 (text, duration_s, error)。"""
    with tempfile.NamedTemporaryFile(suffix='.json', prefix='asr_', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        result = subprocess.run(
            ['bl', 'speech', 'recognize', '--url', filepath,
             '--out', tmp_path, '--output', 'json'],
            capture_output=True, text=True, timeout=ASR_TIMEOUT,
        )
        if result.returncode != 0:
            err = result.stderr.strip() or result.stdout.strip()
            return None, None, err[:500] if err else f'exit {result.returncode}'

        with open(tmp_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        props = data.get('properties', {})
        duration_ms = props.get('original_duration_in_milliseconds')
        duration_s = duration_ms / 1000.0 if duration_ms else None

        transcripts = data.get('transcripts', [])
        if not transcripts:
            return None, duration_s, 'ASR 返回空结果（文件中可能没有语音）'

        texts = [t.get('text', '') for t in transcripts]
        full_text = '\n'.join(texts).strip()
        if not full_text:
            return None, duration_s, 'ASR 返回空文本'

        return full_text, duration_s, None

    except subprocess.TimeoutExpired:
        return None, None, f'ASR 超时 (>{ASR_TIMEOUT}s)'
    except json.JSONDecodeError as e:
        return None, None, f'JSON 解析失败: {e}'
    except Exception as e:
        return None, None, f'未知错误: {e}'
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# ═══════════════════════════════════════════════════
# 输出
# ═══════════════════════════════════════════════════

def save_xlsx(results: list[dict], output_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ASR转写结果'

    hf = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, name in enumerate(FIELD_NAMES, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = halign

    for row_idx, entry in enumerate(results, 2):
        for col_idx, name in enumerate(FIELD_NAMES, 1):
            value = entry[name]
            if name == '时长(秒)' and value is not None:
                value = round(value, 2)
            elif name == '文件大小(MB)' and value is not None:
                value = round(value, 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if name == '处理状态':
                cell.font = Font(color='006100' if value == '成功' else 'FF0000', bold=True)

    widths = {'文件名': 42, '文件路径': 64, '时长(秒)': 11, '文件大小(MB)': 13,
              '完整文本': 72, '处理状态': 10, '处理耗时(秒)': 13, '错误信息': 44}
    for col, name in enumerate(FIELD_NAMES, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 14)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(FIELD_NAMES))}{len(results) + 1}'
    wb.save(output_path)


def save_csv(results: list[dict], output_path: str) -> None:
    import csv
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_NAMES, extrasaction='ignore')
        writer.writeheader()
        for entry in results:
            row = dict(entry)
            if row.get('时长(秒)') is not None:
                row['时长(秒)'] = round(row['时长(秒)'], 2)
            if row.get('文件大小(MB)') is not None:
                row['文件大小(MB)'] = round(row['文件大小(MB)'], 2)
            writer.writerow(row)


# ═══════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════

def process_folder(folder_path: str, output_format: str = 'xlsx',
                   output_path: str | None = None) -> None:
    folder = Path(folder_path).resolve()
    if not folder.is_dir():
        print(f'❌ 错误：{folder_path} 不是有效目录')
        sys.exit(1)

    video_files = sorted([
        f for f in folder.iterdir()
        if f.is_file() and f.suffix.lower() in VIDEO_EXTENSIONS
    ])
    if not video_files:
        print(f'❌ 在 {folder} 中没有找到视频文件')
        print(f'   支持的格式: {", ".join(sorted(VIDEO_EXTENSIONS))}')
        sys.exit(1)

    if output_path is None:
        safe_name = (folder.name or 'videos').replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = str(folder / f'ASR转写结果_{safe_name}_{timestamp}.{output_format}')
    else:
        output_path = str(Path(output_path).resolve())

    total = len(video_files)
    print(f'\n📁 {folder}')
    print(f'🎬 {total} 个视频  |  📄 {output_format.upper()} → {output_path}')
    print(f'{"─" * 60}\n')

    results, overall_start = [], time.time()

    for idx, filepath in enumerate(video_files, 1):
        print(f'[{idx}/{total}] {filepath.name}', end=' ', flush=True)

        entry = dict.fromkeys(FIELD_NAMES, '')
        entry.update({'文件名': filepath.name, '文件路径': str(filepath),
                       '时长(秒)': None, '文件大小(MB)': None,
                       '完整文本': '', '处理状态': '处理中',
                       '处理耗时(秒)': None, '错误信息': ''})

        t0 = time.time()
        entry['文件大小(MB)'] = round(get_file_size_mb(str(filepath)), 2)

        text, duration_s, error = run_asr(str(filepath))
        entry['处理耗时(秒)'] = round(time.time() - t0, 1)
        entry['时长(秒)'] = round(duration_s, 2) if duration_s else None

        if text:
            entry['完整文本'] = text
            entry['处理状态'] = '成功'
            print(f'✅ {entry["处理耗时(秒)"]}s')
        else:
            entry['处理状态'] = '失败'
            entry['错误信息'] = error or '未知错误'
            print(f'❌ {(error or "未知错误")[:80]}')

        results.append(entry)

        # 每完成一个文件立即落盘，防止中途欠费/断网/崩溃导致已处理结果丢失
        if output_format == 'xlsx':
            save_xlsx(results, output_path)
        else:
            save_csv(results, output_path)

    success = sum(1 for r in results if r['处理状态'] == '成功')
    total_dur = sum((r['时长(秒)'] or 0) for r in results if r['处理状态'] == '成功')

    print(f'\n{"═" * 60}')
    print(f'✅ {success}  /  ❌ {total - success}  /  📦 {total}')
    print(f'⏱ {round(time.time() - overall_start, 1)}s  |  🎵 音频总长 {total_dur:.0f}s')
    print(f'📄 {output_path}')
    print(f'{"═" * 60}')


def main():
    parser = argparse.ArgumentParser(
        description='批量视频 ASR 转写 — 阿里云百炼 fun-asr')
    parser.add_argument('--folder', '-f', required=True)
    parser.add_argument('--format', '-fmt', choices=['xlsx', 'csv'], default='xlsx')
    parser.add_argument('--output', '-o', default=None)
    args = parser.parse_args()
    process_folder(args.folder, args.format, args.output)


if __name__ == '__main__':
    main()
